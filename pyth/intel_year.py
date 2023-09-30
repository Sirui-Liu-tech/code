from openpyxl import load_workbook



wb=load_workbook(r"C:\Users\86189\Desktop\code\pyth\Intel_UPE_ComparisonChart_1.xlsx")

sheet=wb[wb.sheetnames[1]]


for i in range(2,2661):
    b=sheet["B"+str(i)].value
    print(b)
    if b:
        sheet["B"+str(i)]=b[3:]

wb.save(r"C:\Users\86189\Desktop\code\pyth\Intel_UPE_ComparisonChart_1.xlsx")
