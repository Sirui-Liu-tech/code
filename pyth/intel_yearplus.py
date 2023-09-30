from openpyxl import load_workbook



wb=load_workbook(r"C:\Users\86189\Desktop\code\pyth\Intel_UPE_ComparisonChart_1.xlsx")

sheet=wb[wb.sheetnames[1]]


for i in range(2,200):
    b=sheet["B"+str(i)].value
    print(b)
    if b:
        if int(b)>50:
            sheet["B"+str(i)]="19"+str(b)
        elif b:
            sheet["B"+str(i)]="20"+str(b)
    print(sheet["B"+str(i)].value)

wb.save(r"C:\Users\86189\Desktop\code\pyth\Intel_UPE_ComparisonChart_1.xlsx")